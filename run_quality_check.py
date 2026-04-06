#!/usr/bin/env python3
"""
GRAIAN CAPITAL MANAGEMENT
Data Quality Pipeline

This tool doesn't need to know your data schema in advance.
Drop any files in — it discovers the structure, finds the issues,
cleans the data, and tells you exactly what to fix in Power BI.

USAGE:
    python run_quality_check.py

WHAT THIS DOES:
    1. Loads all data sources defined in config.json
    2. Detects schema drift vs the previous run (dataset swap guard)
    3. Discovers schema automatically — any file, any columns
    4. Runs data quality checks across all tables
    5. Separates actionable issues from informational observations
    6. Applies remediations (forward-fill, outlier flags)
    7. Exports cleaned files — same names, same structure as input
    8. Exports quality_checklist.xlsx — review and action in Power BI
    9. Exports data_dictionary.xlsx — auto-generated governance doc
    10. Exports quality_report.pdf — stakeholder summary

OUTPUTS:
    output/cleaned_<your_file>.xlsx   → original sheet names preserved
    output/cleaned_<your_file>.csv    → for each CSV source
    output/quality_checklist.xlsx     → Power Query transformation guide
    output/data_dictionary.xlsx       → auto-generated data dictionary
    output/quality_report.pdf         → stakeholder report
    output/schema_baseline.json       → schema snapshot for drift detection

REQUIREMENTS:
    pip install -r requirements.txt
"""

import json
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

from data_quality import (
    SchemaProfiler, DataQualityChecker, Reporter, SuggestionEngine
)
from data_quality.governance import SchemaDriftDetector, DataDictionaryExporter

# ─────────────────────────────────────────────────────────────────────────────
# STEP 0 — Configuration
# ─────────────────────────────────────────────────────────────────────────────

CONFIG_PATH = Path("config.json")

if not CONFIG_PATH.exists():
    print("❌  config.json not found.")
    sys.exit(1)

with open(CONFIG_PATH, encoding="utf-8") as f:
    CONFIG = json.load(f)

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

BASELINE_PATH = OUTPUT_DIR / "schema_baseline.json"


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Ingestion
# ─────────────────────────────────────────────────────────────────────────────

def load_tables(config: dict) -> dict:
    tables  = {}
    sources = config.get("sources", {})

    for name, src in sources.items():
        file_path = Path(src["file"])
        if not file_path.exists():
            print(f"  ⚠️  File not found: {file_path} — skipping '{name}'")
            continue

        if src["type"] == "excel":
            df = pd.read_excel(file_path, sheet_name=src.get("sheet", 0))
        elif src["type"] == "csv":
            df = pd.read_csv(file_path, sep=src.get("delimiter", ","))
            df.columns = df.columns.str.strip().str.lstrip("\ufeff")
        else:
            print(
                f"  ⚠️  Unknown source type '{src['type']}' "
                f"— skipping '{name}'")
            continue

        for col in src.get("date_columns", []):
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce")

        tables[name] = df

    return tables


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Remediation
# ─────────────────────────────────────────────────────────────────────────────

def apply_remediation(tables: dict, profiles: dict) -> dict:
    cleaned = {k: v.copy() for k, v in tables.items()}

    for name, df in cleaned.items():
        profile = profiles.get(name)
        if not profile:
            continue

        for col, cp in profile.columns.items():
            if col not in df.columns:
                continue

            if cp.inferred_type == "numeric" and cp.null_count > 0:
                date_cols = [
                    c for c, p in profile.columns.items()
                    if p.inferred_type == "date"
                ]
                if date_cols:
                    df = df.sort_values(date_cols[0])
                    df[col] = df[col].ffill()

            if cp.inferred_type == "numeric" and cp.outlier_count > 0:
                flag_col = f"{col}-quality-flag"
                lo, hi   = cp.outlier_fence_lo, cp.outlier_fence_hi
                series   = pd.to_numeric(df[col], errors="coerce")
                df[flag_col] = np.where(
                    (series < lo) | (series > hi), "OUTLIER", "OK"
                )

        cleaned[name] = df

    return cleaned


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Export
# ─────────────────────────────────────────────────────────────────────────────

def export_outputs(cleaned: dict, issues: list, suggestions: list) -> list:
    paths   = []
    sources = CONFIG.get("sources", {})

    file_groups = {}
    for name, src in sources.items():
        if name not in cleaned:
            continue
        fname = Path(src["file"]).name
        ftype = src.get("type")
        sheet = src.get("sheet")

        if ftype == "excel" and sheet:
            file_groups.setdefault(fname, {})[sheet] = cleaned[name]
        else:
            p = OUTPUT_DIR / f"cleaned_{fname}"
            cleaned[name].to_csv(p, index=False)
            paths.append(str(p))

    for fname, sheets in file_groups.items():
        out_path = OUTPUT_DIR / f"cleaned_{fname}"
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        paths.append(str(out_path))

    # ── quality checklist ─────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb     = Workbook()
    ws     = wb.active
    ws.title = "Quality Checklist"
    thin   = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cs(row, col, value, bold=False, bg=None, wrap=False, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Arial", size=9, bold=bold)
        c.alignment = Alignment(
            horizontal=align, vertical="top", wrap_text=wrap)
        c.border = border
        if bg:
            c.fill = PatternFill("solid", start_color=bg)

    headers = [
        "#", "Done", "Severity", "Category", "Table", "Column",
        "Issue", "Suggestion", "Power Query Step", "DAX Hint"
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color="1C2B4A")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
    ws.row_dimensions[1].height = 22

    sev_bg  = {
        "HIGH":   "FFF0F0",
        "MEDIUM": "FFFBEA",
        "LOW":    "F2FBF2",
    }
    sug_map = {s["description"]: s for s in suggestions}

    # FIX 5: only actionable issues in the checklist
    actionable_issues = [i for i in issues if not i.is_observation]

    for i, issue in enumerate(actionable_issues, 1):
        r   = i + 1
        bg  = sev_bg.get(issue.severity, "FFFFFF")
        sug = sug_map.get(issue.description, {})
        cs(r, 1, i,                              align="center", bg=bg)
        cs(r, 2, "☐ Open", align="center",       bg=bg)
        cs(r, 3, issue.severity, bold=True,      align="center", bg=bg)
        cs(r, 4, issue.category,                 bg=bg)
        cs(r, 5, issue.table,                    bg=bg)
        cs(r, 6, issue.column or "—",            bg=bg)
        cs(r, 7, issue.description, wrap=True,   bg=bg)
        cs(r, 8, issue.suggestion,  wrap=True,   bg=bg)
        cs(r, 9, sug.get("powerquery_hint", ""), wrap=True, bg=bg)
        cs(r, 10, sug.get("dax_hint", ""),       wrap=True, bg=bg)
        ws.row_dimensions[r].height = 60

    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"☐ Open,☑ Done"',
        allow_blank=False,
        showDropDown=False)
    ws.add_data_validation(dv)
    dv.sqref = f"B2:B{len(actionable_issues) + 1}"

    ws.auto_filter.ref        = (
        f"A1:{get_column_letter(10)}{len(actionable_issues) + 1}")
    ws.freeze_panes           = "A2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows       = "1:1"

    for col, w in enumerate([4, 6, 9, 13, 14, 16, 45, 45, 45, 40], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    cl_path = OUTPUT_DIR / "quality_checklist.xlsx"
    wb.save(cl_path)
    paths.append(str(cl_path))

    # Note: PDF path is appended in main() after generate_pdf(), not here
    return paths


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():

    # ── Governance: schema drift check ───────────────────────────────────────
    drift_detector = SchemaDriftDetector(BASELINE_PATH)

    # ── Agent 1: Ingest ───────────────────────────────────────────────────────
    tables = load_tables(CONFIG)

    # ── Schema drift: compare against saved baseline ──────────────────────────
    drift_events = drift_detector.check(tables, CONFIG)
    # Save current schema as new baseline (always update after a successful run)
    drift_detector.save(tables, CONFIG)

    # ── Agent 2: Discover + Check ─────────────────────────────────────────────
    profiler             = SchemaProfiler(tables)
    profiles, discovered = profiler.run()
    checker              = DataQualityChecker(tables, CONFIG, profiles, discovered)
    all_findings         = checker.run()

    # FIX 5: separate issues from observations
    issues       = [f for f in all_findings if not f.is_observation]
    observations = [f for f in all_findings if f.is_observation]

    engine      = SuggestionEngine(issues, CONFIG)
    suggestions = engine.generate()

    # ── Reporter ──────────────────────────────────────────────────────────────
    reporter = Reporter(CONFIG, issues, observations, suggestions, tables)
    reporter.print_header()
    reporter.print_ingestion()

    if drift_events:
        reporter.print_drift(drift_events)

    reporter.print_issues()
    reporter.print_observations()

    # ── Agent 3: Remediate ────────────────────────────────────────────────────
    cleaned = apply_remediation(tables, profiles)
    reporter.print_remediation()

    # ── Agent 4: Export ───────────────────────────────────────────────────────
    output_paths = export_outputs(cleaned, all_findings, suggestions)

    # ── Governance: data dictionary ───────────────────────────────────────────
    dict_exporter = DataDictionaryExporter(profiles)
    dict_path     = dict_exporter.export(OUTPUT_DIR / "data_dictionary.xlsx")
    output_paths.append(str(dict_path))

    reporter.print_suggestions()
    reporter.print_outputs(output_paths)

    # ── HTML (multi-page) ─────────────────────────────────────────────────────
    bl_info = drift_detector.get_baseline_info()
    if bl_info:
        bl_files = ", ".join(bl_info.get("files", [])) or "unknown files"
        bl_date  = bl_info.get("saved_at", "unknown date")
        bl_rows  = f"{bl_info.get('total_rows', 0):,}"
        baseline_date = f"{bl_files} — saved {bl_date} ({bl_rows} rows)"
    else:
        baseline_date = "No previous run found — this is the first run"
    reporter.generate_html(
        str(OUTPUT_DIR / "quality_report.html"),
        drift_events=drift_events,
        baseline_date=baseline_date)
    output_paths.append(str(OUTPUT_DIR / "quality_report.html"))

    # ── PDF ───────────────────────────────────────────────────────────────────
    # generate_html sets reporter._pdf_filename — use the same name so the
    # file on disk matches the download button in the HTML report
    pdf_name = getattr(reporter, "_pdf_filename", "quality_report.pdf")
    pdf_path = OUTPUT_DIR / pdf_name
    reporter.generate_pdf(str(pdf_path))
    output_paths.append(str(pdf_path))


if __name__ == "__main__":
    main()