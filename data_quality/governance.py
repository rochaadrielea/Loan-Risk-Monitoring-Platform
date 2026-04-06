# data_quality/governance.py
# Governance Layer — schema drift detection + data dictionary export
# Author: Adriele Rocha Weisz

from __future__ import annotations
import json
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import numpy as np


# ─────────────────────────────────────────────────────────────────────────────
# SCHEMA DRIFT DETECTOR
#
# The baseline is a JSON file saved to disk after every run.
# It survives: Flask restarts, computer reboots, closing the browser.
# It persists until you manually delete output/schema_baseline.json.
#
# What the baseline stores per table:
#   - columns: list of column names
#   - row_count: total rows
#   - files: list of source filenames that contributed to this table
#   - value_sets: for categorical/id columns — the set of unique values
#   - numeric_ranges: for numeric columns — {col: {min, max, mean}}
#   - saved_at: ISO timestamp of when this baseline was created
# ─────────────────────────────────────────────────────────────────────────────

class SchemaDriftDetector:

    # Thresholds
    ROW_COUNT_DRIFT_PCT   = 20    # flag if row count changes > 20%
    NUMERIC_RANGE_DRIFT_PCT = 50  # flag if min/max changes > 50%

    def __init__(self, baseline_path: Path):
        self.baseline_path = Path(baseline_path)

    # ── public API ────────────────────────────────────────────────────────────

    def check(self, tables: dict, config: dict = None) -> list[dict]:
        """
        Compare current tables against the saved baseline.
        Returns list of drift events. Empty list = no drift.
        First run (no baseline yet) returns [].
        """
        if not self.baseline_path.exists():
            return []
        baseline = self._load()
        current  = self._snapshot(tables, config)
        return self._diff(baseline, current)

    def save(self, tables: dict, config: dict = None):
        """
        Save current schema as the new baseline.
        Called after every successful pipeline run.
        Persists to disk — survives restarts.
        """
        snapshot = self._snapshot(tables, config)
        snapshot["_saved_at"] = datetime.now().isoformat()
        self.baseline_path.write_text(
            json.dumps(snapshot, indent=2, default=str),
            encoding="utf-8")

    def get_baseline_info(self) -> dict:
        """
        Returns metadata about the saved baseline for display.
        { saved_at, files, tables, total_rows }
        """
        if not self.baseline_path.exists():
            return {}
        try:
            data = json.loads(
                self.baseline_path.read_text(encoding="utf-8"))
            saved_at = data.get("_saved_at", "")[:19].replace("T", " ")
            # Collect all source filenames mentioned across all tables
            files = set()
            total_rows = 0
            table_names = []
            for key, val in data.items():
                if key.startswith("_"):
                    continue
                table_names.append(key)
                total_rows += val.get("row_count", 0)
                for f in val.get("files", []):
                    files.add(f)
            return {
                "saved_at":   saved_at,
                "files":      sorted(files),
                "tables":     table_names,
                "total_rows": total_rows,
            }
        except Exception:
            return {}

    # ── snapshot ──────────────────────────────────────────────────────────────

    def _snapshot(self, tables: dict, config: dict = None) -> dict:
        """
        Build a rich snapshot of the current dataset.
        Stores: columns, row_count, source files, value sets, numeric ranges.
        """
        sources = (config or {}).get("sources", {})
        snap = {}

        for name, df in tables.items():
            src = sources.get(name, {})
            fname = Path(src.get("file", "")).name if src.get("file") else name

            # Value sets for categorical/id columns (≤ 50 unique values)
            value_sets = {}
            numeric_ranges = {}

            for col in df.columns:
                series = df[col].dropna()
                if series.empty:
                    continue

                # Try numeric
                nums = pd.to_numeric(series, errors="coerce").dropna()
                if len(nums) > 0 and len(nums) / max(len(series), 1) > 0.8:
                    numeric_ranges[col] = {
                        "min":  round(float(nums.min()), 6),
                        "max":  round(float(nums.max()), 6),
                        "mean": round(float(nums.mean()), 6),
                    }
                else:
                    # Categorical — save value set if small enough
                    n_unique = series.nunique()
                    if n_unique <= 50:
                        value_sets[col] = sorted(
                            str(v) for v in series.unique())

            snap[name] = {
                "columns":        list(df.columns),
                "row_count":      len(df),
                "files":          [fname] if fname else [],
                "value_sets":     value_sets,
                "numeric_ranges": numeric_ranges,
            }

        return snap

    # ── load baseline ─────────────────────────────────────────────────────────

    def _load(self) -> dict:
        data = json.loads(
            self.baseline_path.read_text(encoding="utf-8"))
        # Strip metadata keys — leave only table entries
        return {k: v for k, v in data.items()
                if not k.startswith("_")}

    # ── diff ──────────────────────────────────────────────────────────────────

    def _diff(self, baseline: dict, current: dict) -> list[dict]:
        events = []

        b_tables = set(baseline.keys())
        c_tables = set(current.keys())

        # ── Tables added / removed ────────────────────────────────────────────
        for t in c_tables - b_tables:
            events.append({
                "type":   "TABLE_ADDED",
                "table":  t,
                "risk":   "LOW",
                "detail": (
                    f"Table [{t}] is new — it was not present in the "
                    f"previous run. Verify it is correctly mapped in "
                    f"Power Query and that no existing DAX measure needs updating."
                ),
            })

        for t in b_tables - c_tables:
            events.append({
                "type":   "TABLE_REMOVED",
                "table":  t,
                "risk":   "HIGH",
                "detail": (
                    f"Table [{t}] was present before but is missing now. "
                    f"Every DAX measure and Power Query step that references "
                    f"[{t}] will fail immediately when Power BI refreshes."
                ),
            })

        # ── Per-table comparison ──────────────────────────────────────────────
        for t in b_tables & c_tables:
            b = baseline[t]
            c = current[t]

            old_cols = set(b.get("columns", []))
            new_cols = set(c.get("columns", []))

            # Columns added
            for col in sorted(new_cols - old_cols):
                events.append({
                    "type":   "COLUMN_ADDED",
                    "table":  t,
                    "risk":   "LOW",
                    "detail": (
                        f"[{t}.{col}] is a new column not seen before. "
                        f"Verify it is mapped correctly in Power Query. "
                        f"If any DAX measure should use this column, update it now."
                    ),
                })

            # Columns removed — highest risk
            for col in sorted(old_cols - new_cols):
                events.append({
                    "type":   "COLUMN_REMOVED",
                    "table":  t,
                    "risk":   "HIGH",
                    "detail": (
                        f"[{t}.{col}] existed before but is gone now. "
                        f"Any DAX measure or Power Query step that references "
                        f"[{col}] will break immediately. Check all measures "
                        f"in the transactions, products, and fx tables."
                    ),
                })

            # Row count change
            old_rows = b.get("row_count", 0)
            new_rows = c.get("row_count", 0)
            if old_rows > 0:
                delta_pct = abs(new_rows - old_rows) / old_rows * 100
                if delta_pct > self.ROW_COUNT_DRIFT_PCT:
                    direction = "increased" if new_rows > old_rows else "decreased"
                    events.append({
                        "type":   "ROW_COUNT_CHANGE",
                        "table":  t,
                        "risk":   "MEDIUM",
                        "detail": (
                            f"[{t}] had {old_rows:,} rows before. "
                            f"Now it has {new_rows:,} rows — "
                            f"a {direction} of {delta_pct:.0f}%. "
                            f"This may mean the new dataset covers a different "
                            f"time period, or that rows were added or removed. "
                            f"Verify the dataset is complete before refreshing Power BI."
                        ),
                    })

            # Value set drift — new category values not seen before
            old_vs = b.get("value_sets", {})
            new_vs = c.get("value_sets", {})
            for col in set(old_vs.keys()) & set(new_vs.keys()):
                old_vals = set(old_vs[col])
                new_vals = set(new_vs[col])
                added_vals   = new_vals - old_vals
                removed_vals = old_vals - new_vals
                if added_vals:
                    events.append({
                        "type":   "VALUE_SET_CHANGED",
                        "table":  t,
                        "risk":   "MEDIUM",
                        "detail": (
                            f"[{t}.{col}] has {len(added_vals)} new value(s) "
                            f"not seen before: {sorted(added_vals)[:5]}. "
                            f"These may be new portfolio codes, product types, "
                            f"or issuer codes. Verify they are correctly handled "
                            f"in Power BI filters and slicers."
                        ),
                    })
                if removed_vals:
                    events.append({
                        "type":   "VALUE_SET_CHANGED",
                        "table":  t,
                        "risk":   "LOW",
                        "detail": (
                            f"[{t}.{col}] has {len(removed_vals)} value(s) "
                            f"that existed before but are now missing: "
                            f"{sorted(removed_vals)[:5]}. "
                            f"These codes may have been renamed, merged, or "
                            f"are simply not present in this dataset's date range."
                        ),
                    })

            # Numeric range drift — min/max changed significantly
            old_nr = b.get("numeric_ranges", {})
            new_nr = c.get("numeric_ranges", {})
            for col in set(old_nr.keys()) & set(new_nr.keys()):
                old_min = old_nr[col].get("min", 0)
                old_max = old_nr[col].get("max", 0)
                new_min = new_nr[col].get("min", 0)
                new_max = new_nr[col].get("max", 0)

                # Max changed significantly
                if old_max != 0:
                    max_delta = abs(new_max - old_max) / abs(old_max) * 100
                    if max_delta > self.NUMERIC_RANGE_DRIFT_PCT:
                        direction = "increased" if new_max > old_max else "decreased"
                        events.append({
                            "type":   "NUMERIC_RANGE_CHANGED",
                            "table":  t,
                            "risk":   "HIGH" if max_delta > 200 else "MEDIUM",
                            "detail": (
                                f"[{t}.{col}] maximum value {direction} "
                                f"from {old_max:,.4f} to {new_max:,.4f} "
                                f"({max_delta:.0f}% change). "
                                f"This is the kind of change that catches "
                                f"data entry errors — for example, "
                                f"loan-property-value-ratio going from 2.15 to 33.33. "
                                f"Review the highest values in this column before loading."
                            ),
                        })

                # Min changed significantly
                if old_min != 0:
                    min_delta = abs(new_min - old_min) / abs(old_min) * 100
                    if min_delta > self.NUMERIC_RANGE_DRIFT_PCT:
                        direction = "increased" if new_min > old_min else "decreased"
                        events.append({
                            "type":   "NUMERIC_RANGE_CHANGED",
                            "table":  t,
                            "risk":   "MEDIUM",
                            "detail": (
                                f"[{t}.{col}] minimum value {direction} "
                                f"from {old_min:,.4f} to {new_min:,.4f} "
                                f"({min_delta:.0f}% change). "
                                f"Review the lowest values in this column "
                                f"before loading into Power BI."
                            ),
                        })

        # Sort by risk: HIGH first, then MEDIUM, then LOW
        risk_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
        events.sort(key=lambda e: risk_order.get(e.get("risk", "LOW"), 3))
        return events


# ─────────────────────────────────────────────────────────────────────────────
# DATA DICTIONARY EXPORTER (unchanged)
# ─────────────────────────────────────────────────────────────────────────────

class DataDictionaryExporter:

    def __init__(self, profiles: dict):
        self.profiles = profiles

    def export(self, output_path: Path) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("  ⚠️  openpyxl not available — skipping data dictionary.")
            return output_path

        wb = Workbook()
        ws = wb.active
        ws.title = "Data Dictionary"

        NAVY = "1C2B4A"; GOLD = "B89650"; LIGHT = "F5F6F8"
        thin   = Side(style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def cell(row, col, value, bold=False, bg=None,
                 align="left", wrap=False, color=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(name="Arial", size=9, bold=bold,
                          color=color or "1A1A1A")
            c.alignment = Alignment(horizontal=align, vertical="top",
                                    wrap_text=wrap)
            c.border = border
            if bg:
                c.fill = PatternFill("solid", start_color=bg)

        headers = ["Table","Column","Inferred Type","Nullable","Null %",
                   "Unique Values","Unique %","Min / Date Min","Max / Date Max",
                   "Sample Values","Notes"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", start_color=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.row_dimensions[1].height = 22

        row_idx = 2
        for tname, profile in self.profiles.items():
            for cname, cp in profile.columns.items():
                bg = LIGHT if row_idx % 2 == 0 else "FFFFFF"
                notes = []
                if cp.is_likely_pk:     notes.append("inferred PK")
                if cp.is_likely_fk:     notes.append("inferred FK")
                if cp.outlier_count > 0: notes.append(f"{cp.outlier_count} outlier(s)")
                if cp.cardinality_flag == "HIGH": notes.append("high cardinality")
                if cp.has_negatives and cp.inferred_type == "numeric":
                    notes.append("contains negatives")

                if cp.inferred_type == "numeric":
                    vmin = f"{cp.num_min:,.4f}" if cp.num_min is not None else "—"
                    vmax = f"{cp.num_max:,.4f}" if cp.num_max is not None else "—"
                elif cp.inferred_type == "date":
                    vmin = str(cp.date_min.date()) if cp.date_min else "—"
                    vmax = str(cp.date_max.date()) if cp.date_max else "—"
                else:
                    vmin = vmax = "—"

                sample = ", ".join(str(v) for v in cp.sample_values[:3])
                cell(row_idx, 1,  tname,                    bg=bg)
                cell(row_idx, 2,  cname,                    bg=bg, bold=True)
                cell(row_idx, 3,  cp.inferred_type,         bg=bg, align="center")
                cell(row_idx, 4,  "Yes" if cp.null_count > 0 else "No",
                                                             bg=bg, align="center")
                cell(row_idx, 5,  f"{cp.null_pct:.1f}%",   bg=bg, align="center")
                cell(row_idx, 6,  cp.unique_count,          bg=bg, align="center")
                cell(row_idx, 7,  f"{cp.unique_pct:.1f}%", bg=bg, align="center")
                cell(row_idx, 8,  vmin,                     bg=bg)
                cell(row_idx, 9,  vmax,                     bg=bg)
                cell(row_idx, 10, sample, wrap=True,        bg=bg)
                cell(row_idx, 11, "; ".join(notes) if notes else "",
                     wrap=True, bg=bg, color=GOLD if notes else None)
                ws.row_dimensions[row_idx].height = 18
                row_idx += 1

        widths = [16, 22, 14, 9, 8, 13, 10, 16, 16, 30, 30]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.freeze_panes    = "A2"

        ws_meta = wb.create_sheet("Metadata")
        ws_meta["A1"] = "Generated"
        ws_meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_meta["A2"] = "Tables"
        ws_meta["B2"] = len(self.profiles)
        ws_meta["A3"] = "Total Columns"
        ws_meta["B3"] = sum(len(p.columns) for p in self.profiles.values())
        ws_meta["A4"] = "Tool"
        ws_meta["B4"] = "Graian Capital Management — Data Quality Pipeline"

        wb.save(output_path)
        return output_path